import os
import sys
import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime
from openpyxl import load_workbook

EXCEL_NAME = "BASE_INVENTARIO_LISTA.xlsx"
SHEET_INV = "Hoja1"
SHEET_MOV = "Movimientos"

def resource_dir():
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

EXCEL_PATH = os.path.join(resource_dir(), EXCEL_NAME)

class InventarioApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Control de Inventario")
        self.root.geometry("900x600")
        self.root.minsize(820, 540)

        self.modo = tk.StringVar(value="SALIDA")
        self.cantidad = tk.IntVar(value=1)

        self.crear_ui()
        self.verificar_excel()
        self.entry_codigo.focus_set()

    def crear_ui(self):
        top = ttk.Frame(self.root, padding=12)
        top.pack(fill="x")

        ttk.Label(top, text="CONTROL DE INVENTARIO",
                  font=("Segoe UI", 16, "bold")).pack(side="left")

        self.lbl_excel = ttk.Label(top, text="")
        self.lbl_excel.pack(side="right")

        modo_frame = ttk.LabelFrame(self.root, text="Tipo de movimiento", padding=12)
        modo_frame.pack(fill="x", padx=12, pady=(0, 8))

        ttk.Radiobutton(modo_frame, text="ENTRADA  (+)",
                        variable=self.modo, value="ENTRADA").pack(side="left", padx=10)
        ttk.Radiobutton(modo_frame, text="SALIDA  (-)",
                        variable=self.modo, value="SALIDA").pack(side="left", padx=10)

        ttk.Label(modo_frame, text="Cantidad por escaneo:").pack(side="left", padx=(30, 6))
        ttk.Spinbox(modo_frame, from_=1, to=99999,
                    textvariable=self.cantidad, width=8).pack(side="left")

        scan = ttk.LabelFrame(self.root, text="Lector de código de barras", padding=14)
        scan.pack(fill="x", padx=12, pady=8)

        ttk.Label(scan, text="Escanea el código del producto (Modelo):").pack(anchor="w")
        self.entry_codigo = ttk.Entry(scan, font=("Segoe UI", 20))
        self.entry_codigo.pack(fill="x", pady=8)
        self.entry_codigo.bind("<Return>", self.procesar)

        self.lbl_estado = ttk.Label(scan, text="Listo para escanear.",
                                    font=("Segoe UI", 11, "bold"))
        self.lbl_estado.pack(anchor="w")

        info = ttk.LabelFrame(self.root, text="Producto", padding=10)
        info.pack(fill="x", padx=12, pady=8)

        self.lbl_modelo = ttk.Label(info, text="Modelo: —")
        self.lbl_modelo.grid(row=0, column=0, sticky="w", padx=5, pady=3)
        self.lbl_marca = ttk.Label(info, text="Marca: —")
        self.lbl_marca.grid(row=0, column=1, sticky="w", padx=25, pady=3)
        self.lbl_desc = ttk.Label(info, text="Descripción: —")
        self.lbl_desc.grid(row=1, column=0, columnspan=2, sticky="w", padx=5, pady=3)
        self.lbl_stock = ttk.Label(info, text="Existencia: —",
                                   font=("Segoe UI", 12, "bold"))
        self.lbl_stock.grid(row=2, column=0, sticky="w", padx=5, pady=6)

        tabla_frame = ttk.LabelFrame(self.root, text="Últimos movimientos", padding=8)
        tabla_frame.pack(fill="both", expand=True, padx=12, pady=8)

        cols = ("hora", "modelo", "marca", "mov", "cant", "stock")
        self.tabla = ttk.Treeview(tabla_frame, columns=cols, show="headings")
        labels = {
            "hora": "Hora", "modelo": "Modelo", "marca": "Marca",
            "mov": "Movimiento", "cant": "Cantidad", "stock": "Existencia"
        }
        widths = {"hora": 90, "modelo": 180, "marca": 150, "mov": 100, "cant": 80, "stock": 90}

        for c in cols:
            self.tabla.heading(c, text=labels[c])
            self.tabla.column(c, width=widths[c], anchor="center")

        self.tabla.pack(fill="both", expand=True)

        bottom = ttk.Frame(self.root, padding=12)
        bottom.pack(fill="x")
        ttk.Button(bottom, text="Recargar", command=self.verificar_excel).pack(side="left")
        ttk.Button(bottom, text="Abrir carpeta", command=self.abrir_carpeta).pack(side="left", padx=8)
        ttk.Label(bottom, text="El archivo Excel debe permanecer cerrado mientras se escanea.").pack(side="right")

    def verificar_excel(self):
        if not os.path.exists(EXCEL_PATH):
            self.lbl_excel.config(text=f"No encontrado: {EXCEL_NAME}")
            self.lbl_estado.config(text=f"Coloca {EXCEL_NAME} junto al programa.")
            return False
        try:
            wb = load_workbook(EXCEL_PATH, read_only=True)
            ok = SHEET_INV in wb.sheetnames and SHEET_MOV in wb.sheetnames
            wb.close()
            if not ok:
                raise ValueError("El archivo no tiene las hojas esperadas.")
            self.lbl_excel.config(text=EXCEL_NAME)
            self.lbl_estado.config(text="Excel cargado. Listo para escanear.")
            return True
        except Exception as e:
            self.lbl_estado.config(text=f"Error al leer Excel: {e}")
            return False

    def buscar_modelo(self, ws, codigo):
        codigo = str(codigo).strip().upper()
        for fila in range(2, ws.max_row + 1):
            valor = ws.cell(fila, 1).value
            if valor is not None and str(valor).strip().upper() == codigo:
                return fila
        return None

    def procesar(self, event=None):
        codigo = self.entry_codigo.get().strip()
        self.entry_codigo.delete(0, tk.END)

        if not codigo:
            return
        if not self.verificar_excel():
            self.root.bell()
            self.entry_codigo.focus_set()
            return

        try:
            cant = int(self.cantidad.get())
            if cant <= 0:
                raise ValueError
        except Exception:
            messagebox.showwarning("Cantidad", "La cantidad debe ser un entero mayor que cero.")
            self.entry_codigo.focus_set()
            return

        try:
            wb = load_workbook(EXCEL_PATH)
            ws = wb[SHEET_INV]
            mov = wb[SHEET_MOV]

            fila = self.buscar_modelo(ws, codigo)
            if fila is None:
                wb.close()
                self.lbl_estado.config(text=f"No encontrado: {codigo}")
                self.lbl_modelo.config(text=f"Modelo: {codigo}")
                self.lbl_marca.config(text="Marca: —")
                self.lbl_desc.config(text="Descripción: —")
                self.lbl_stock.config(text="Existencia: —")
                self.root.bell()
                self.entry_codigo.focus_set()
                return

            modelo = ws.cell(fila, 1).value or ""
            marca = ws.cell(fila, 2).value or ""
            desc = ws.cell(fila, 3).value or ""
            stock = ws.cell(fila, 4).value

            try:
                stock = int(stock or 0)
            except Exception:
                stock = 0

            if self.modo.get() == "ENTRADA":
                nuevo = stock + cant
            else:
                nuevo = stock - cant
                if nuevo < 0:
                    wb.close()
                    messagebox.showwarning(
                        "Existencia insuficiente",
                        f"{modelo}\nExistencia actual: {stock}\nSalida solicitada: {cant}"
                    )
                    self.lbl_estado.config(text="Movimiento cancelado por existencia insuficiente.")
                    self.root.bell()
                    self.entry_codigo.focus_set()
                    return

            ws.cell(fila, 4).value = nuevo

            ahora = datetime.now()
            mov.append([
                ahora.strftime("%Y-%m-%d"),
                ahora.strftime("%H:%M:%S"),
                modelo,
                marca,
                desc,
                self.modo.get(),
                cant,
                nuevo
            ])

            wb.save(EXCEL_PATH)
            wb.close()

            self.lbl_modelo.config(text=f"Modelo: {modelo}")
            self.lbl_marca.config(text=f"Marca: {marca}")
            self.lbl_desc.config(text=f"Descripción: {desc}")
            self.lbl_stock.config(text=f"Existencia: {nuevo}")
            signo = "+" if self.modo.get() == "ENTRADA" else "-"
            self.lbl_estado.config(text=f"{self.modo.get()} registrada: {signo}{cant} | Existencia actual: {nuevo}")

            self.tabla.insert("", 0, values=(
                ahora.strftime("%H:%M:%S"), modelo, marca,
                self.modo.get(), cant, nuevo
            ))

            items = self.tabla.get_children()
            if len(items) > 100:
                self.tabla.delete(items[-1])

        except PermissionError:
            messagebox.showerror(
                "Excel está abierto",
                f"Cierra {EXCEL_NAME} en Excel y vuelve a escanear."
            )
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo registrar el movimiento:\n{e}")

        self.entry_codigo.focus_set()

    def abrir_carpeta(self):
        try:
            os.startfile(resource_dir())
        except Exception:
            pass

if __name__ == "__main__":
    root = tk.Tk()
    app = InventarioApp(root)
    root.mainloop()
